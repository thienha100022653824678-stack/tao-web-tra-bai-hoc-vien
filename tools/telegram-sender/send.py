import os
import sys
import time
import random
import asyncio
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

from telethon import TelegramClient
from telethon.sessions import StringSession
from telethon.errors import FloodWaitError

# Load env variables from root directory
dotenv_path = os.path.join(os.path.dirname(__file__), '../../.env')
load_dotenv(dotenv_path=dotenv_path)

def get_value(row, col_idx):
    if col_idx == -1 or col_idx >= len(row):
        return ""
    val = row[col_idx]
    return str(val).strip() if val is not None else ""

def normalize_channel_id(channel_id_raw):
    channel_id_raw = str(channel_id_raw).strip()
    if not channel_id_raw:
        return None

    # If it is numeric (e.g. 1002265382 or -1002265382)
    if channel_id_raw.startswith("-100"):
        try:
            return int(channel_id_raw)
        except ValueError:
            return channel_id_raw

    if channel_id_raw.startswith("-"):
        try:
            return int(channel_id_raw)
        except ValueError:
            return channel_id_raw

    try:
        return int("-100" + channel_id_raw)
    except ValueError:
        # Username case: e.g. @channelname
        return channel_id_raw

async def main():
    print("=" * 60)
    print(" CÔNG CỤ GỬI TIN NHẮN BÀI HỌC TELEGRAM (TELETHON SENDER)")
    print("=" * 60)

    # Check for dry-run mode
    is_dry_run = '--dry-run' in sys.argv or os.getenv('DRY_RUN') == 'true'
    if is_dry_run:
        print("⚠️ CHẾ ĐỘ CHẠY THỬ (DRY RUN): KHÔNG GỬI TIN NHẮN THẬT.")
        print("-" * 60)

    # File paths
    input_file = os.path.join(os.path.dirname(__file__), '../../result_links.xlsx')
    output_file = os.path.join(os.path.dirname(__file__), '../../sent_report.xlsx')

    if not os.path.exists(input_file):
        print(f"Lỗi: Không tìm thấy file kết quả '{input_file}' ở thư mục gốc.")
        print("Vui lòng chạy import trước (hoặc kéo file 'result_links.xlsx' vào đây).")
        return

    # Load result links
    try:
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        
        # Read header columns
        header_row = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        
        required_cols = ['original_channel_name', 'normalized_title', 'telegram_chat_id', 'post_id', 'student_url', 'message_text']
        col_indices = {}
        for col in required_cols:
            if col not in header_row:
                print(f"Lỗi: File Excel '{input_file}' thiếu cột bắt buộc: '{col}'")
                return
            col_indices[col] = header_row.index(col)

        # Find or add 'send_status' column
        if 'send_status' in header_row:
            status_idx = header_row.index('send_status')
        else:
            status_idx = len(header_row)
            header_row.append('send_status')
            sheet.cell(row=1, column=status_idx + 1, value='send_status')

        # Add optional error/notes columns if not present
        if 'sent_at' not in header_row:
            sent_at_idx = len(header_row)
            header_row.append('sent_at')
            sheet.cell(row=1, column=sent_at_idx + 1, value='sent_at')
        else:
            sent_at_idx = header_row.index('sent_at')

        if 'error_log' not in header_row:
            error_idx = len(header_row)
            header_row.append('error_log')
            sheet.cell(row=1, column=error_idx + 1, value='error_log')
        else:
            error_idx = header_row.index('error_log')

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        return

    # Load API Credentials
    api_id_raw = os.getenv('TELEGRAM_API_ID')
    api_hash = os.getenv('TELEGRAM_API_HASH')
    telegram_session = os.getenv('TELEGRAM_SESSION')

    if not api_id_raw or not api_hash:
        print("Lỗi: Thiếu TELEGRAM_API_ID hoặc TELEGRAM_API_HASH trong file .env.")
        return

    try:
        api_id = int(api_id_raw)
    except ValueError:
        print("Lỗi: TELEGRAM_API_ID phải là một số nguyên.")
        return

    # Initialize Telegram client
    print("Đang khởi tạo Telegram Client...")
    if telegram_session:
        # Load portable StringSession
        client = TelegramClient(StringSession(telegram_session), api_id, api_hash)
    else:
        # Fallback to local session file
        session_file = os.path.join(os.path.dirname(__file__), 'telegram_session')
        client = TelegramClient(session_file, api_id, api_hash)

    try:
        await client.start()
        print("Đăng nhập Telegram thành công.")
    except Exception as e:
        print(f"Lỗi khi đăng nhập Telegram: {e}")
        print("Vui lòng chạy `python tools/telegram-sender/login.py` để tạo session trước.")
        return

    # Process and send messages
    sent_count = 0
    skipped_count = 0
    failed_count = 0
    
    rows_to_process = list(sheet.iter_rows(min_row=2))
    print(f"Tìm thấy {len(rows_to_process)} dòng trong file. Bắt đầu xử lý...")

    for index, row in enumerate(rows_to_process):
        excel_row_num = index + 2
        
        row_values = [cell.value for cell in row]
        status = get_value(row_values, status_idx)
        
        # Skip if already successfully sent
        if status == 'SENT':
            print(f"Dòng {excel_row_num}: Đã gửi thành công ở phiên trước. Bỏ qua.")
            skipped_count += 1
            continue

        title = get_value(row_values, col_indices['normalized_title'])
        chat_id_raw = get_value(row_values, col_indices['telegram_chat_id'])
        message_text = get_value(row_values, col_indices['message_text'])
        channel_name = get_value(row_values, col_indices['original_channel_name']) or title

        if not chat_id_raw or not message_text:
            print(f"Dòng {excel_row_num}: Bỏ qua do thiếu Telegram Chat ID hoặc nội dung tin nhắn.")
            sheet.cell(row=excel_row_num, column=status_idx + 1, value='SKIPPED')
            sheet.cell(row=excel_row_num, column=error_idx + 1, value='Thiếu ID chat hoặc tin nhắn')
            skipped_count += 1
            continue

        chat_id = normalize_channel_id(chat_id_raw)
        
        if is_dry_run:
            # DRY RUN MODE
            print(f"[{index+1}/{len(rows_to_process)}] [DRY RUN] Sẽ gửi đến {chat_id_raw} ({channel_name}):")
            print(f"--- Tin nhắn ---\n{message_text}\n----------------")
            sheet.cell(row=excel_row_num, column=status_idx + 1, value='SKIPPED')
            sheet.cell(row=excel_row_num, column=sent_at_idx + 1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            sheet.cell(row=excel_row_num, column=error_idx + 1, value='Chạy thử (Dry run)')
            skipped_count += 1
            continue

        # LIVE MODE
        try:
            print(f"[{index+1}/{len(rows_to_process)}] Đang gửi tới: {channel_name} ({chat_id_raw})...")
            
            await client.send_message(
                chat_id,
                message_text,
                link_preview=False
            )
            
            # Log success
            sheet.cell(row=excel_row_num, column=status_idx + 1, value='SENT')
            sheet.cell(row=excel_row_num, column=sent_at_idx + 1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            sheet.cell(row=excel_row_num, column=error_idx + 1, value='')
            
            sent_count += 1
            print(f" -> ✅ THÀNH CÔNG!")
            
            # Save progress incrementally to avoid data loss if interrupted
            wb.save(output_file)

            # Random delay 2-5 seconds as per YÊU CẦU 4
            delay = random.uniform(2.0, 5.0)
            await asyncio.sleep(delay)

        except FloodWaitError as e:
            wait_time = int(e.seconds) + 5
            print(f" -> ⏳ Bị giới hạn Flood. Yêu cầu chờ {wait_time} giây...")
            sheet.cell(row=excel_row_num, column=status_idx + 1, value='FAILED')
            sheet.cell(row=excel_row_num, column=error_idx + 1, value=f'FloodWait: Yêu cầu chờ {wait_time}s')
            wb.save(output_file)
            
            await asyncio.sleep(wait_time)
            # Re-append current index to retry? No, let's keep running and continue.

        except Exception as e:
            failed_count += 1
            error_str = str(e)
            print(f" -> ❌ THẤT BẠI: {error_str}")
            
            sheet.cell(row=excel_row_num, column=status_idx + 1, value='FAILED')
            sheet.cell(row=excel_row_num, column=error_idx + 1, value=error_str[:500])
            
            wb.save(output_file)
            
            # Delay even on error to be safe
            await asyncio.sleep(2.0)

    # Disconnect client
    await client.disconnect()

    # Save final report workbook
    wb.save(output_file)
    print("\n" + "=" * 60)
    print(" HOÀN THÀNH PHIÊN GỬI TIN NHẮN")
    print("=" * 60)
    print(f" - Đã gửi mới: {sent_count}")
    print(f" - Bỏ qua / Chạy thử: {skipped_count}")
    print(f" - Thất bại / Lỗi: {failed_count}")
    print(f" - File báo cáo đã xuất: '{output_file}'")
    print("=" * 60 + "\n")

if __name__ == '__main__':
    asyncio.run(main())
