import os
import requests
import openpyxl
from dotenv import load_dotenv

# Load env variables from root directory
dotenv_path = os.path.join(os.path.dirname(__file__), '../../.env')
load_dotenv(dotenv_path=dotenv_path)

def main():
    print("=" * 60)
    print(" CÔNG CỤ NHẬP BÀI HỌC CỤC BỘ (LOCAL POST IMPORT CLI)")
    print("=" * 60)

    supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
    student_app_url = os.getenv('NEXT_PUBLIC_STUDENT_APP_URL') or 'https://www.yeunauan.live'
    
    if not student_app_url.startswith('http'):
        # Sanitize Vercel env formatting if it has '='
        if '=' in student_app_url:
            student_app_url = student_app_url.split('=')[-1]
    student_app_url = student_app_url.strip().rstrip('/')

    if not supabase_url or not supabase_key:
        print("Lỗi: Không tìm thấy NEXT_PUBLIC_SUPABASE_URL hoặc SUPABASE_SERVICE_ROLE_KEY trong file .env")
        return

    # Check for input file in root
    input_file = os.path.join(os.path.dirname(__file__), '../../import_posts.xlsx')
    output_file = os.path.join(os.path.dirname(__file__), '../../result_links.xlsx')

    if not os.path.exists(input_file):
        print(f"Lỗi: Không tìm thấy file '{input_file}' ở thư mục gốc.")
        print("Vui lòng sao chép file 'import_posts.xlsx' vào thư mục gốc của dự án.")
        return

    print("Đang đọc file 'import_posts.xlsx'...")
    try:
        wb = openpyxl.load_workbook(input_file, read_only=True)
        sheet = wb.active
        
        # Read header columns
        header_row = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        
        required_cols = ['original_channel_name', 'normalized_title', 'telegram_chat_id', 'recipe', 'match_status']
        col_indices = {}
        for col in required_cols:
            if col not in header_row:
                print(f"Lỗi: File Excel thiếu cột bắt buộc: '{col}'")
                return
            col_indices[col] = header_row.index(col)
        
        # Optional notes column index
        notes_idx = header_row.index('notes') if 'notes' in header_row else -1

        posts_to_insert = []
        original_rows_data = [] # To map return Supabase UUIDs back to Excel values
        
        # Read data rows
        row_num = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or all(v is None for v in row):
                continue
                
            match_status = str(row[col_indices['match_status']] or '').strip().upper()
            if match_status != 'MATCHED':
                continue
                
            title = str(row[col_indices['normalized_title']] or '').strip()
            recipe = str(row[col_indices['recipe']] or '').strip()
            chat_id = str(row[col_indices['telegram_chat_id']] or '').strip()
            channel_name = str(row[col_indices['original_channel_name']] or '').strip()
            notes = str(row[notes_idx]) if notes_idx != -1 and row[notes_idx] is not None else ''

            if not title or not recipe:
                print(f"Bỏ qua dòng {row_num}: Thiếu Tiêu đề (normalized_title) hoặc Công thức (recipe).")
                continue

            posts_to_insert.append({
                "title": title,
                "recipe": recipe,
                "images": [],
                "views": 0,
                "telegram_chat_id": chat_id if chat_id else None,
                "original_channel_name": channel_name if channel_name else None
            })
            
            original_rows_data.append({
                "original_channel_name": channel_name,
                "normalized_title": title,
                "telegram_chat_id": chat_id
            })

        if not posts_to_insert:
            print("Không tìm thấy dòng nào có trạng thái 'MATCHED' để import.")
            return

        print(f"Tìm thấy {len(posts_to_insert)} bài học sẵn sàng import.")
        print("Đang tiến hành bulk insert vào Supabase...")

        # Supabase REST API endpoint for posts table insertion
        supabase_url = supabase_url.strip().rstrip('/')
        url = f"{supabase_url}/rest/v1/posts"
        headers = {
            "apikey": supabase_key.strip(),
            "Authorization": f"Bearer {supabase_key.strip()}",
            "Content-Type": "application/json",
            "Prefer": "return=representation"
        }

        response = requests.post(url, json=posts_to_insert, headers=headers)
        
        if response.status_code not in [200, 201]:
            print(f"Lỗi kết nối database (HTTP {response.status_code}): {response.text}")
            return
            
        inserted_posts = response.json()
        print(f"Đã import thành công {len(inserted_posts)} bài học lên database.")

        # Match database generated IDs with the Excel data and output result links
        results_wb = openpyxl.Workbook()
        results_sheet = results_wb.active
        results_sheet.title = "Result Links"
        
        # Headers matching YÊU CẦU 2
        results_sheet.append([
            'original_channel_name',
            'normalized_title',
            'telegram_chat_id',
            'post_id',
            'student_url',
            'message_text',
            'send_status'
        ])

        for i, post in enumerate(inserted_posts):
            post_id = post.get('id')
            title = post.get('title')
            chat_id = post.get('telegram_chat_id') or ''
            channel_name = post.get('original_channel_name') or ''
            
            student_url = f"{student_app_url}/post/{post_id}"
            
            # Format message_text as per YÊU CẦU 3
            message_text = f"📚 Bài học của lớp: {title}\n\nBạn xem nội dung bài học tại đây:\n{student_url}"
            
            results_sheet.append([
                channel_name,
                title,
                chat_id,
                post_id,
                student_url,
                message_text,
                '' # send_status is empty initially
            ])

        results_wb.save(output_file)
        print(f"Đã xuất tệp kết quả: '{output_file}'")
        print("Bây giờ bạn đã sẵn sàng chạy công cụ gửi Telegram bằng lệnh:")
        print("  npm run telegram:send")
        print("-" * 60)

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")

if __name__ == '__main__':
    main()
